#!/usr/bin/env python3
"""
generate_examples.py
====================
Generates all example Excel data files for the LabQC project.

These files contain realistic QC, validation, and sigma metric data
for common molecular diagnostic assays (SARS-CoV-2, HIV-1, HBV, HCV, TB/MTB)
and selected clinical chemistry analytes.

Usage:
    py -3.12 example_sets/generate_examples.py

Requires: openpyxl
"""

import os
import random
import math
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SEED = 42
random.seed(SEED)

BASE_DIR = Path(__file__).resolve().parent
QC_DIR = BASE_DIR / "qc_data"
VAL_DIR = BASE_DIR / "validation_data"
SIGMA_DIR = BASE_DIR / "sigma_data"

# Ensure directories exist
for d in [QC_DIR, VAL_DIR, SIGMA_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# Styling helpers
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, num_cols):
    """Apply formatting to header row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_width(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def rand_ct(mean, sd):
    """Generate a random Ct value with normal distribution, rounded to 2 dp."""
    return round(random.gauss(mean, sd), 2)


def well_position(index, cols=12):
    """Convert 0-based index to a microplate well position (A01..H12)."""
    row_letter = chr(65 + index // cols)  # A-H
    col_num = (index % cols) + 1
    return f"{row_letter}{col_num:02d}"


def make_date_series(start_str, n, gap_days=1):
    """Generate a list of date strings starting from start_str."""
    start = datetime.strptime(start_str, "%Y-%m-%d")
    return [(start + timedelta(days=i * gap_days)).strftime("%Y-%m-%d") for i in range(n)]


# ===================================================================
# 1. QuantStudio SARS-CoV-2 E-gene  (20 runs, L1 + L2)
# ===================================================================
def generate_quantstudio_sars_cov2():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Well", "Well Position", "Sample Name", "Target Name",
        "Task", "Reporter", "Quencher", "CT", "Ct Mean", "Ct SD",
        "Quantity", "Quantity Mean", "Quantity SD",
        "Automatic Ct Threshold", "Ct Threshold", "Baseline Start", "Baseline End",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "SARS-CoV-2 E-gene"
    dates = make_date_series("2025-06-01", 20, gap_days=1)
    well_idx = 0

    # Controls definition
    controls = [
        ("QC-L1 Low Positive", 32.0, 0.8),
        ("QC-L2 High Positive", 22.5, 0.4),
        ("NTC", None, None),  # Negative / No Template Control
    ]

    for run_i in range(20):
        for ctrl_name, ct_mean, ct_sd in controls:
            well_idx += 1
            wp = well_position(well_idx - 1)

            if ct_mean is None:
                ct_val = "Undetermined"
                ct_mean_val = ""
                ct_sd_val = ""
                task = "NTC"
            else:
                ct_val = rand_ct(ct_mean, ct_sd)
                ct_mean_val = round(ct_mean + random.gauss(0, 0.05), 2)
                ct_sd_val = round(abs(random.gauss(ct_sd, 0.05)), 3)
                task = "STANDARD"

            ws.append([
                well_idx,
                wp,
                f"{ctrl_name} Run{run_i + 1:02d} ({dates[run_i]})",
                target,
                task,
                "FAM",
                "NFQ-MGB",
                ct_val,
                ct_mean_val if ct_mean_val != "" else "",
                ct_sd_val if ct_sd_val != "" else "",
                "",  # Quantity
                "",  # Quantity Mean
                "",  # Quantity SD
                "TRUE",
                0.2,
                3,
                15,
            ])

    auto_width(ws)
    path = QC_DIR / "quantstudio_sars_cov2.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 2. QuantStudio HIV-1 Viral Load  (15 runs, L1 + L2 + L3)
# ===================================================================
def generate_quantstudio_hiv1():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Well", "Well Position", "Sample Name", "Target Name",
        "Task", "Reporter", "Quencher", "CT", "Ct Mean", "Ct SD",
        "Quantity", "Quantity Mean", "Quantity SD",
        "Automatic Ct Threshold", "Ct Threshold", "Baseline Start", "Baseline End",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "HIV-1 RNA"
    dates = make_date_series("2025-07-01", 15, gap_days=2)
    well_idx = 0

    controls = [
        ("QC-L1 Low", 35.0, 1.0),
        ("QC-L2 Medium", 28.0, 0.6),
        ("QC-L3 High", 20.0, 0.3),
        ("NTC", None, None),
    ]

    for run_i in range(15):
        for ctrl_name, ct_mean, ct_sd in controls:
            well_idx += 1
            wp = well_position(well_idx - 1)

            if ct_mean is None:
                ct_val = "Undetermined"
                ct_mean_val = ""
                ct_sd_val = ""
                task = "NTC"
            else:
                ct_val = rand_ct(ct_mean, ct_sd)
                ct_mean_val = round(ct_mean + random.gauss(0, 0.05), 2)
                ct_sd_val = round(abs(random.gauss(ct_sd, 0.05)), 3)
                task = "STANDARD"

            ws.append([
                well_idx,
                wp,
                f"{ctrl_name} Run{run_i + 1:02d} ({dates[run_i]})",
                target,
                task,
                "FAM",
                "NFQ-MGB",
                ct_val,
                ct_mean_val if ct_mean_val != "" else "",
                ct_sd_val if ct_sd_val != "" else "",
                "", "", "",
                "TRUE", 0.15, 3, 15,
            ])

    auto_width(ws)
    path = QC_DIR / "quantstudio_hiv1.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 3. Bio-Rad CFX HBV DNA  (10 runs, L1 + L2)
# ===================================================================
def generate_cfx_hbv():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quantification Cq Results"

    headers = ["Well", "Fluor", "Target", "Content", "Sample", "Cq"]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "HBV DNA"
    dates = make_date_series("2025-08-01", 10, gap_days=3)
    well_idx = 0

    controls = [
        ("Pos Ctrl", "QC-L1 Low Positive", 33.5, 0.9),
        ("Pos Ctrl", "QC-L2 High Positive", 24.0, 0.5),
        ("NTC", "NTC", None, None),
    ]

    for run_i in range(10):
        for content, sample_name, ct_mean, ct_sd in controls:
            well_idx += 1
            wp = well_position(well_idx - 1)

            if ct_mean is None:
                cq = "N/A"
            else:
                cq = rand_ct(ct_mean, ct_sd)

            ws.append([
                wp,
                "FAM",
                target,
                content,
                f"{sample_name} Run{run_i + 1:02d} ({dates[run_i]})",
                cq,
            ])

    auto_width(ws)
    path = QC_DIR / "cfx_hbv.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 4. Generic TB/MTB  (12 runs, L1 + L2)
# ===================================================================
def generate_generic_tb_mtb():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = ["Sample ID", "Target", "Control Level", "Ct Value", "Mean", "SD",
               "Run Date", "Instrument", "Operator"]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "MTB IS6110"
    dates = make_date_series("2025-09-01", 12, gap_days=2)
    operators = ["Tech-A", "Tech-B", "Tech-C"]

    controls = [
        ("L1", "Low Positive", 30.0, 0.7),
        ("L2", "High Positive", 18.5, 0.4),
    ]

    sample_counter = 0
    for run_i in range(12):
        for level, level_name, ct_mean, ct_sd in controls:
            sample_counter += 1
            ct_val = rand_ct(ct_mean, ct_sd)
            ws.append([
                f"TB-QC-{sample_counter:04d}",
                target,
                f"{level} ({level_name})",
                ct_val,
                ct_mean,
                ct_sd,
                dates[run_i],
                "GeneXpert Infinity-48s",
                random.choice(operators),
            ])

    # Negative controls
    for run_i in range(12):
        sample_counter += 1
        ws.append([
            f"TB-QC-{sample_counter:04d}",
            target,
            "NTC (Negative)",
            "Undetermined",
            "",
            "",
            dates[run_i],
            "GeneXpert Infinity-48s",
            random.choice(operators),
        ])

    auto_width(ws)
    path = QC_DIR / "generic_tb_mtb.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 5. QC With Deliberate Westgard Violations
# ===================================================================
def generate_qc_with_violations():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Well", "Well Position", "Sample Name", "Target Name",
        "Task", "Reporter", "Quencher", "CT", "Ct Mean", "Ct SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "SARS-CoV-2 E-gene"
    dates = make_date_series("2025-10-01", 20, gap_days=1)

    # L1: mean=32.0, SD=0.8 ; L2: mean=22.5, SD=0.4
    l1_mean, l1_sd = 32.0, 0.8
    l2_mean, l2_sd = 22.5, 0.4
    well_idx = 0

    for run_i in range(20):
        run_num = run_i + 1

        # --- L1 ---
        if run_num == 8:
            # 1-3s violation: value > mean + 3*SD
            l1_ct = round(l1_mean + 3.5 * l1_sd, 2)  # 34.80
        elif run_num in (15, 16, 17, 18):
            # 4-1s violation: 4 consecutive > mean + 1*SD (same side)
            l1_ct = round(l1_mean + 1.3 * l1_sd, 2)  # 33.04
        else:
            l1_ct = rand_ct(l1_mean, l1_sd)

        well_idx += 1
        ws.append([
            well_idx, well_position(well_idx - 1),
            f"QC-L1 Low Positive Run{run_num:02d} ({dates[run_i]})",
            target, "STANDARD", "FAM", "NFQ-MGB",
            l1_ct, l1_mean, l1_sd,
        ])

        # --- L2 ---
        if run_num in (12, 13):
            # 2-2s violation: 2 consecutive > mean + 2*SD
            l2_ct = round(l2_mean + 2.3 * l2_sd, 2)  # 23.42
        else:
            l2_ct = rand_ct(l2_mean, l2_sd)

        well_idx += 1
        ws.append([
            well_idx, well_position(well_idx - 1),
            f"QC-L2 High Positive Run{run_num:02d} ({dates[run_i]})",
            target, "STANDARD", "FAM", "NFQ-MGB",
            l2_ct, l2_mean, l2_sd,
        ])

        # --- NTC ---
        well_idx += 1
        ws.append([
            well_idx, well_position(well_idx - 1),
            f"NTC Run{run_num:02d} ({dates[run_i]})",
            target, "NTC", "FAM", "NFQ-MGB",
            "Undetermined", "", "",
        ])

    # Add a "Violations Legend" sheet for reference
    ws2 = wb.create_sheet("Violations Legend")
    ws2.append(["Run", "Control", "Expected Violation", "Description"])
    style_header(ws2, 4)
    ws2.append([8, "L1", "1-3s", "Single value exceeds mean + 3*SD (Ct = 34.80)"])
    ws2.append([12, "L2", "2-2s (start)", "Two consecutive values exceed mean + 2*SD"])
    ws2.append([13, "L2", "2-2s (end)", "Continuation of 2-2s violation (Ct = 23.42)"])
    ws2.append([15, "L1", "4-1s (start)", "Four consecutive values exceed mean + 1*SD"])
    ws2.append([16, "L1", "4-1s", "Continuation of 4-1s violation"])
    ws2.append([17, "L1", "4-1s", "Continuation of 4-1s violation"])
    ws2.append([18, "L1", "4-1s (end)", "End of 4-1s violation (Ct = 33.04)"])
    auto_width(ws2)

    auto_width(ws)
    path = QC_DIR / "qc_with_violations.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 6. QC With Control Lot Change
# ===================================================================
def generate_qc_lot_change():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Well", "Well Position", "Sample Name", "Target Name",
        "Task", "Reporter", "Quencher", "CT", "Ct Mean", "Ct SD",
        "Control Lot", "Lot Expiry",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "SARS-CoV-2 RdRp"
    dates = make_date_series("2025-04-01", 20, gap_days=1)
    well_idx = 0

    for run_i in range(20):
        run_num = run_i + 1
        well_idx += 1

        if run_num <= 10:
            lot = "CTL-2024-087"
            lot_expiry = "2025-06-30"
            ct_mean, ct_sd = 25.0, 0.5
        else:
            lot = "CTL-2025-012"
            lot_expiry = "2026-03-31"
            ct_mean, ct_sd = 24.8, 0.45

        ct_val = rand_ct(ct_mean, ct_sd)

        ws.append([
            well_idx, well_position(well_idx - 1),
            f"QC-L2 Run{run_num:02d} ({dates[run_i]})",
            target, "STANDARD", "FAM", "NFQ-MGB",
            ct_val, ct_mean, ct_sd,
            lot, lot_expiry,
        ])

    # Add a metadata sheet
    ws2 = wb.create_sheet("Lot Change Info")
    ws2.append(["Parameter", "Old Lot", "New Lot"])
    style_header(ws2, 3)
    ws2.append(["Lot Number", "CTL-2024-087", "CTL-2025-012"])
    ws2.append(["Manufacturer", "AccuPlex SARS-CoV-2 (SeraCare)", "AccuPlex SARS-CoV-2 (SeraCare)"])
    ws2.append(["Runs", "1-10", "11-20"])
    ws2.append(["Expected Ct Mean", 25.0, 24.8])
    ws2.append(["Expected Ct SD", 0.5, 0.45])
    ws2.append(["Expiry", "2025-06-30", "2026-03-31"])
    ws2.append(["Lot Change Date", "2025-04-11", ""])
    ws2.append(["Reason", "Lot expiration approaching", ""])
    auto_width(ws2)

    auto_width(ws)
    path = QC_DIR / "qc_lot_change.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 7. LOD Study - SARS-CoV-2
# ===================================================================
def generate_lod_sars_cov2():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LOD Study"

    headers = [
        "Sample ID", "Target", "Concentration (copies/mL)",
        "Replicate", "Ct Value", "Detection (Pos/Neg)",
        "Run Date", "Operator",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "SARS-CoV-2 E-gene"

    # (concentration, ct_mean, ct_sd, detection_rate)
    levels = [
        (0, None, None, 0.0),         # Blank
        (50, 37.0, 1.5, 0.55),        # Near LOD - ~55% detection
        (100, 35.0, 1.0, 0.80),       # ~80% detection
        (200, 33.0, 0.8, 0.95),       # LOD target - ~95% detection
        (500, 30.5, 0.5, 1.0),        # Above LOD
        (1000, 28.5, 0.4, 1.0),       # Well above LOD
    ]

    sample_counter = 0
    for conc, ct_mean, ct_sd, det_rate in levels:
        for rep in range(1, 21):
            sample_counter += 1
            detected = random.random() < det_rate

            if conc == 0 or not detected:
                ct_val = "Undetermined"
                det_status = "Neg"
            else:
                ct_val = rand_ct(ct_mean, ct_sd)
                # Ensure Ct stays reasonable
                if ct_val > 40:
                    ct_val = "Undetermined"
                    det_status = "Neg"
                else:
                    det_status = "Pos"

            ws.append([
                f"LOD-{sample_counter:04d}",
                target,
                conc,
                rep,
                ct_val,
                det_status,
                "2025-11-15",
                "Tech-A",
            ])

    # Summary sheet
    ws2 = wb.create_sheet("LOD Summary")
    ws2.append(["Concentration (copies/mL)", "N Replicates", "N Detected",
                "Detection Rate (%)", "Mean Ct", "SD Ct"])
    style_header(ws2, 6)
    ws2.append([0, 20, 0, "0.0", "N/A", "N/A"])
    ws2.append([50, 20, "~11", "~55", "~37.0", "~1.5"])
    ws2.append([100, 20, "~16", "~80", "~35.0", "~1.0"])
    ws2.append(["200 (Claimed LOD)", 20, "~19", "~95", "~33.0", "~0.8"])
    ws2.append([500, 20, 20, "100", "~30.5", "~0.5"])
    ws2.append([1000, 20, 20, "100", "~28.5", "~0.4"])
    auto_width(ws2)

    auto_width(ws)
    path = VAL_DIR / "lod_sars_cov2.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 8. LOQ Study - HIV-1 Viral Load
# ===================================================================
def generate_loq_hiv1():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LOQ Study"

    headers = [
        "Sample ID", "Target", "Nominal Concentration (copies/mL)",
        "Log10 Nominal", "Replicate", "Ct Value",
        "Measured Concentration (copies/mL)", "Log10 Measured",
        "Run Date", "Operator",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "HIV-1 RNA"

    # (conc, ct_mean, ct_sd)
    # CV decreases with increasing concentration (realistic behavior)
    levels = [
        (100,   35.5, 1.2),   # CV ~3.4%
        (500,   31.0, 0.8),   # CV ~2.6%
        (1000,  28.5, 0.6),   # CV ~2.1%
        (2500,  26.0, 0.45),  # CV ~1.7%
        (5000,  24.0, 0.35),  # CV ~1.5%
        (10000, 22.0, 0.25),  # CV ~1.1%
    ]

    # Simple Ct-to-concentration conversion: log10(conc) = (40 - Ct) / slope
    # slope ~ 3.3 (typical for 100% efficiency)
    slope = 3.322  # log2(10) for perfect efficiency
    intercept = 40.0

    sample_counter = 0
    for conc, ct_mean, ct_sd in levels:
        log10_nominal = round(math.log10(conc), 3)
        for rep in range(1, 11):
            sample_counter += 1
            ct_val = rand_ct(ct_mean, ct_sd)

            # Back-calculate measured concentration from Ct
            log10_meas = round((intercept - ct_val) / slope, 3)
            meas_conc = round(10 ** log10_meas, 1)

            ws.append([
                f"LOQ-{sample_counter:04d}",
                target,
                conc,
                log10_nominal,
                rep,
                ct_val,
                meas_conc,
                log10_meas,
                "2025-12-01",
                "Tech-B",
            ])

    auto_width(ws)
    path = VAL_DIR / "loq_hiv1.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 9. Intra-Run Precision - HBV
# ===================================================================
def generate_precision_intra_hbv():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Intra-Run Precision"

    headers = [
        "Sample ID", "Target", "Control Level", "Well Position",
        "Replicate", "Ct Value", "Run Date", "Instrument", "Operator",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "HBV DNA"
    ct_mean, ct_sd = 24.0, 0.5

    for rep in range(1, 21):
        ct_val = rand_ct(ct_mean, ct_sd)
        ws.append([
            f"HBV-PREC-{rep:03d}",
            target,
            "L2 (High Positive)",
            well_position(rep - 1),
            rep,
            ct_val,
            "2025-12-10",
            "Bio-Rad CFX96 Touch",
            "Tech-C",
        ])

    # Summary row
    ws.append([])
    ws.append(["Summary Statistics"])
    ws.append(["N Replicates", 20])
    ws.append(["Target Mean Ct", ct_mean])
    ws.append(["Target SD", ct_sd])
    ws.append(["Acceptance: CV%", "<5%"])

    auto_width(ws)
    path = VAL_DIR / "precision_intra_hbv.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 10. Inter-Run Precision - HBV
# ===================================================================
def generate_precision_inter_hbv():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inter-Run Precision"

    headers = [
        "Sample ID", "Target", "Control Level", "Run Number",
        "Run Date", "Replicate", "Ct Value", "Instrument", "Operator",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "HBV DNA"
    ct_mean, ct_sd = 24.0, 0.5
    dates = make_date_series("2025-12-01", 5, gap_days=1)
    operators = ["Tech-A", "Tech-B", "Tech-C", "Tech-A", "Tech-B"]

    sample_counter = 0
    for run_i in range(5):
        # Add slight run-to-run variation
        run_shift = random.gauss(0, 0.15)
        for rep in range(1, 5):
            sample_counter += 1
            ct_val = rand_ct(ct_mean + run_shift, ct_sd)
            ws.append([
                f"HBV-INTER-{sample_counter:03d}",
                target,
                "L2 (High Positive)",
                run_i + 1,
                dates[run_i],
                rep,
                ct_val,
                "Bio-Rad CFX96 Touch",
                operators[run_i],
            ])

    auto_width(ws)
    path = VAL_DIR / "precision_inter_hbv.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 11. Linearity Study - HCV
# ===================================================================
def generate_linearity_hcv():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Linearity Study"

    headers = [
        "Sample ID", "Target", "Nominal Concentration (copies/mL)",
        "Log10 Nominal", "Replicate", "Ct Value",
        "Measured Concentration (copies/mL)", "Log10 Measured",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    target = "HCV RNA"

    # Standard curve: Ct = intercept - slope * log10(conc)
    # Typical: slope ~3.32 (100% efficiency), intercept ~40
    slope = 3.35
    intercept = 40.5

    levels = [100, 500, 1000, 5000, 10000, 50000, 100000]

    sample_counter = 0
    for conc in levels:
        log10_nom = math.log10(conc)
        expected_ct = intercept - slope * log10_nom

        for rep in range(1, 4):  # 3 replicates per level
            sample_counter += 1
            # Add small measurement noise
            noise = random.gauss(0, 0.25)
            ct_val = round(expected_ct + noise, 2)

            # Back-calculate
            log10_meas = round((intercept - ct_val) / slope, 4)
            meas_conc = round(10 ** log10_meas, 1)

            ws.append([
                f"LIN-{sample_counter:03d}",
                target,
                conc,
                round(log10_nom, 4),
                rep,
                ct_val,
                meas_conc,
                log10_meas,
            ])

    # Summary sheet with regression info
    ws2 = wb.create_sheet("Regression Summary")
    ws2.append(["Parameter", "Value"])
    style_header(ws2, 2)
    ws2.append(["Target", target])
    ws2.append(["N Levels", len(levels)])
    ws2.append(["N Total Replicates", sample_counter])
    ws2.append(["Slope", -slope])
    ws2.append(["Intercept", intercept])
    ws2.append(["Efficiency (%)", round((10 ** (1 / slope) - 1) * 100, 1)])
    ws2.append(["R-squared", 0.998])
    ws2.append(["Dynamic Range", f"{levels[0]} - {levels[-1]} copies/mL"])
    ws2.append(["", ""])
    ws2.append(["Note", "R-squared is approximate; recalculate from actual data"])
    auto_width(ws2)

    auto_width(ws)
    path = VAL_DIR / "linearity_hcv.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# 12. Sigma Metric Inputs
# ===================================================================
def generate_sigma_inputs():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sigma Inputs"

    headers = [
        "Assay", "Analyte Category", "TEa%", "Bias%", "CV%",
        "Expected Sigma", "Sigma Rating", "TEa Source",
        "Bias Source", "CV Source",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    assays = [
        ("Glucose", "Clinical Chemistry", 6.96, 1.4, 1.8, 3.09, "Marginal",
         "CLIA/RiliBaK", "EQA scheme 2024", "IQC data 2024-Q4"),
        ("HbA1c", "Clinical Chemistry", 6.0, 1.0, 1.5, 3.33, "Marginal",
         "IFCC/NGSP", "EQA scheme 2024", "IQC data 2024-Q4"),
        ("Creatinine", "Clinical Chemistry", 15.0, 3.2, 2.8, 4.21, "Good",
         "CLIA", "EQA scheme 2024", "IQC data 2024-Q4"),
        ("TSH", "Immunoassay", 23.7, 4.5, 3.2, 6.0, "World Class",
         "RiliBaK/Ricos", "EQA scheme 2024", "IQC data 2024-Q4"),
        ("Troponin I", "Immunoassay", 14.2, 5.1, 4.0, 2.28, "Unacceptable",
         "Ricos/EFLM", "EQA scheme 2024", "IQC data 2024-Q4"),
        ("SARS-CoV-2 Ct", "Molecular", 10.0, 2.0, 1.5, 5.33, "Excellent",
         "Lab-defined (clinical impact)", "Method comparison", "IQC data 2024-Q4"),
        ("HIV-1 VL (log10)", "Molecular", 15.0, 3.0, 2.5, 4.80, "Good",
         "DAIDS/WHO", "VQA proficiency", "IQC data 2024-Q4"),
        ("HBV DNA (log10)", "Molecular", 12.0, 2.5, 2.0, 4.75, "Good",
         "Lab-defined (clinical decision)", "Method comparison", "IQC data 2024-Q4"),
    ]

    for row in assays:
        ws.append(list(row))

    # Add a reference sheet
    ws2 = wb.create_sheet("Sigma Scale Reference")
    ws2.append(["Sigma Value", "Rating", "Defect Rate", "QC Strategy"])
    style_header(ws2, 4)
    ws2.append(["< 2.0", "Unacceptable", "> 308,537 DPM", "Cannot rely on QC alone; redesign method"])
    ws2.append(["2.0 - 3.0", "Poor", "66,807 - 308,537 DPM", "Maximum QC: multi-rule, N=4+"])
    ws2.append(["3.0 - 4.0", "Marginal", "6,210 - 66,807 DPM", "1-3s/2-2s/R-4s, N=2-4"])
    ws2.append(["4.0 - 5.0", "Good", "233 - 6,210 DPM", "1-3s, N=2"])
    ws2.append(["5.0 - 6.0", "Excellent", "3.4 - 233 DPM", "1-3s, N=1-2"])
    ws2.append(["> 6.0", "World Class", "< 3.4 DPM", "Minimal QC needed"])
    auto_width(ws2)

    # TEa sources sheet
    ws3 = wb.create_sheet("TEa Sources")
    ws3.append(["Source", "Description", "URL"])
    style_header(ws3, 3)
    ws3.append(["CLIA", "US Clinical Laboratory Improvement Amendments proficiency testing criteria",
                "https://www.ecfr.gov/current/title-42/chapter-IV/subchapter-G/part-493"])
    ws3.append(["RiliBaK", "German Richtlinie der Bundesarztekammer (Federal Medical Association guidelines)",
                "https://www.bundesaerztekammer.de"])
    ws3.append(["Ricos/EFLM", "Desirable biological variation database (now maintained by EFLM)",
                "https://biologicalvariation.eu"])
    ws3.append(["IFCC/NGSP", "HbA1c standardisation criteria",
                "http://www.ngsp.org"])
    ws3.append(["DAIDS/WHO", "Division of AIDS VQA program for HIV viral load",
                "https://www.hanc.info/labs/virology-quality-assurance.html"])
    ws3.append(["Lab-defined", "TEa set based on clinical decision points and expert consensus",
                "N/A"])
    auto_width(ws3)

    auto_width(ws)
    path = SIGMA_DIR / "sigma_inputs_molecular.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# Main - Generate all files
# ===================================================================
def main():
    print("=" * 60)
    print("LabQC Example Data Generator")
    print("=" * 60)
    print()
    print(f"Output directory: {BASE_DIR}")
    print(f"Random seed: {SEED}")
    print()

    print("[QC Data]")
    generate_quantstudio_sars_cov2()
    generate_quantstudio_hiv1()
    generate_cfx_hbv()
    generate_generic_tb_mtb()
    generate_qc_with_violations()
    generate_qc_lot_change()
    print()

    print("[Validation Data]")
    generate_lod_sars_cov2()
    generate_loq_hiv1()
    generate_precision_intra_hbv()
    generate_precision_inter_hbv()
    generate_linearity_hcv()
    print()

    print("[Sigma Data]")
    generate_sigma_inputs()
    print()

    print("=" * 60)
    print("All files generated successfully.")
    print("=" * 60)


if __name__ == "__main__":
    main()
