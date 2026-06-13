#!/usr/bin/env python3
"""
generate_clinical_chemistry.py
==============================
Generates example Excel data files for non-molecular clinical laboratory
disciplines: clinical chemistry, hematology, immunoassay, plus expanded
sigma metric and validation data.

Westgard rules apply identically to concentration-based QC (chemistry,
hematology, immunoassay) as they do to Ct-based QC (molecular). This
script produces realistic QC control data for each discipline.

Usage:
    py -3.12 example_sets/generate_clinical_chemistry.py

Requires: openpyxl
"""

import os
import random
import math
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SEED = 42
random.seed(SEED)

BASE_DIR = Path(__file__).resolve().parent
CHEM_DIR = BASE_DIR / "clinical_chemistry"
HEMA_DIR = BASE_DIR / "hematology"
IMMU_DIR = BASE_DIR / "immunoassay"
SIGMA_DIR = BASE_DIR / "sigma_data"
VAL_DIR = BASE_DIR / "validation_data"

# Ensure directories exist
for d in [CHEM_DIR, HEMA_DIR, IMMU_DIR, SIGMA_DIR, VAL_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# Styling helpers (matching existing generate_examples.py)
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, num_cols):
    """Apply formatting to header row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_width(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def rand_val(mean, sd):
    """Generate a random value with normal distribution, rounded to appropriate precision."""
    val = random.gauss(mean, sd)
    # Determine decimal places based on magnitude
    if abs(mean) < 0.1:
        return round(val, 4)
    elif abs(mean) < 1:
        return round(val, 3)
    elif abs(mean) < 10:
        return round(val, 2)
    elif abs(mean) < 100:
        return round(val, 1)
    else:
        return round(val, 1)


def make_date_series(start_str, n, gap_days=1):
    """Generate a list of date strings starting from start_str."""
    start = datetime.strptime(start_str, "%Y-%m-%d")
    return [(start + timedelta(days=i * gap_days)).strftime("%Y-%m-%d") for i in range(n)]


# ===================================================================
# CLINICAL CHEMISTRY
# ===================================================================

# -------------------------------------------------------------------
# 1. Glucose QC on Cobas c501
# -------------------------------------------------------------------
def generate_glucose_qc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = [
        "Run Date", "Sample ID", "Analyte", "Level",
        "Value (mg/dL)", "Mean", "SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    analyte = "Glucose"
    dates = make_date_series("2025-09-01", 30, gap_days=1)

    controls = [
        ("L1 (Normal)", 95.0, 2.5),
        ("L2 (Pathological)", 250.0, 6.0),
    ]

    sample_counter = 0
    for run_i in range(30):
        for level_name, mean, sd in controls:
            sample_counter += 1
            value = rand_val(mean, sd)
            ws.append([
                dates[run_i],
                f"GLU-QC-{sample_counter:04d}",
                analyte,
                level_name,
                value,
                mean,
                sd,
            ])

    # Metadata sheet
    ws2 = wb.create_sheet("Assay Info")
    ws2.append(["Parameter", "Value"])
    style_header(ws2, 2)
    ws2.append(["Analyte", "Glucose"])
    ws2.append(["Instrument", "Roche Cobas c501"])
    ws2.append(["Method", "Hexokinase / Glucose-6-phosphate dehydrogenase"])
    ws2.append(["Units", "mg/dL"])
    ws2.append(["TEa (CLIA)", "6.96% or +/- 6 mg/dL (whichever is greater)"])
    ws2.append(["QC Material", "Roche PreciControl ClinChem Multi 1 & 2"])
    ws2.append(["Runs", 30])
    ws2.append(["Levels", "L1 (Normal), L2 (Pathological)"])
    auto_width(ws2)

    auto_width(ws)
    path = CHEM_DIR / "glucose_qc.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# -------------------------------------------------------------------
# 2. Lipid Panel QC on Cobas c501
# -------------------------------------------------------------------
def generate_lipid_panel_qc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = [
        "Run Date", "Sample ID", "Analyte", "Level",
        "Value (mg/dL)", "Mean", "SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    dates = make_date_series("2025-09-01", 20, gap_days=1)

    analytes = [
        ("Total Cholesterol", [("L1 (Normal)", 180.0, 4.5), ("L2 (High)", 280.0, 7.0)]),
        ("HDL Cholesterol", [("L1 (Normal)", 45.0, 1.8), ("L2 (High)", 75.0, 2.5)]),
        ("LDL Cholesterol", [("L1 (Normal)", 100.0, 3.5), ("L2 (High)", 180.0, 5.0)]),
        ("Triglycerides", [("L1 (Normal)", 120.0, 5.0), ("L2 (High)", 300.0, 10.0)]),
    ]

    sample_counter = 0
    for run_i in range(20):
        for analyte_name, levels in analytes:
            for level_name, mean, sd in levels:
                sample_counter += 1
                value = rand_val(mean, sd)
                ws.append([
                    dates[run_i],
                    f"LIP-QC-{sample_counter:04d}",
                    analyte_name,
                    level_name,
                    value,
                    mean,
                    sd,
                ])

    # Metadata sheet
    ws2 = wb.create_sheet("Assay Info")
    ws2.append(["Analyte", "Method", "Units", "TEa Source", "TEa%"])
    style_header(ws2, 5)
    ws2.append(["Total Cholesterol", "Cholesterol oxidase / esterase", "mg/dL", "NCEP", "10.0%"])
    ws2.append(["HDL Cholesterol", "Direct / homogeneous", "mg/dL", "NCEP", "13.0%"])
    ws2.append(["LDL Cholesterol", "Direct / homogeneous", "mg/dL", "NCEP", "12.0%"])
    ws2.append(["Triglycerides", "GPO-PAP", "mg/dL", "NCEP", "15.0%"])
    auto_width(ws2)

    auto_width(ws)
    path = CHEM_DIR / "lipid_panel_qc.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# -------------------------------------------------------------------
# 3. Liver Function QC
# -------------------------------------------------------------------
def generate_liver_function_qc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = [
        "Run Date", "Sample ID", "Analyte", "Level",
        "Value", "Unit", "Mean", "SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    dates = make_date_series("2025-10-01", 25, gap_days=1)

    analytes = [
        ("ALT", "U/L", [("L1 (Normal)", 30.0, 1.5), ("L2 (Pathological)", 120.0, 4.0)]),
        ("AST", "U/L", [("L1 (Normal)", 25.0, 1.2), ("L2 (Pathological)", 100.0, 3.5)]),
        ("ALP", "U/L", [("L1 (Normal)", 70.0, 3.0), ("L2 (Pathological)", 200.0, 6.0)]),
        ("Bilirubin (Total)", "mg/dL", [("L1 (Normal)", 0.8, 0.05), ("L2 (Pathological)", 5.0, 0.15)]),
    ]

    sample_counter = 0
    for run_i in range(25):
        for analyte_name, unit, levels in analytes:
            for level_name, mean, sd in levels:
                sample_counter += 1
                value = rand_val(mean, sd)
                ws.append([
                    dates[run_i],
                    f"LFT-QC-{sample_counter:04d}",
                    analyte_name,
                    level_name,
                    value,
                    unit,
                    mean,
                    sd,
                ])

    # Metadata sheet
    ws2 = wb.create_sheet("Assay Info")
    ws2.append(["Analyte", "Method", "Units", "TEa Source", "TEa%"])
    style_header(ws2, 5)
    ws2.append(["ALT", "IFCC without pyridoxal phosphate", "U/L", "CLIA", "20.0%"])
    ws2.append(["AST", "IFCC without pyridoxal phosphate", "U/L", "CLIA", "20.0%"])
    ws2.append(["ALP", "IFCC / p-Nitrophenyl phosphate", "U/L", "CLIA", "30.0%"])
    ws2.append(["Bilirubin (Total)", "Diazo / Jendrassik-Grof", "mg/dL", "CLIA", "20.0% or +/- 0.4 mg/dL"])
    auto_width(ws2)

    auto_width(ws)
    path = CHEM_DIR / "liver_function_qc.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# -------------------------------------------------------------------
# 4. Renal Function QC
# -------------------------------------------------------------------
def generate_renal_function_qc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = [
        "Run Date", "Sample ID", "Analyte", "Level",
        "Value", "Unit", "Mean", "SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    dates = make_date_series("2025-10-01", 20, gap_days=1)

    analytes = [
        ("Creatinine", "mg/dL", [("L1 (Normal)", 1.0, 0.04), ("L2 (Pathological)", 5.0, 0.12)]),
        ("BUN", "mg/dL", [("L1 (Normal)", 15.0, 0.8), ("L2 (Pathological)", 60.0, 2.0)]),
    ]

    sample_counter = 0
    for run_i in range(20):
        for analyte_name, unit, levels in analytes:
            for level_name, mean, sd in levels:
                sample_counter += 1
                value = rand_val(mean, sd)
                ws.append([
                    dates[run_i],
                    f"RFT-QC-{sample_counter:04d}",
                    analyte_name,
                    level_name,
                    value,
                    unit,
                    mean,
                    sd,
                ])

    # Metadata sheet
    ws2 = wb.create_sheet("Assay Info")
    ws2.append(["Analyte", "Method", "Units", "TEa Source", "TEa%"])
    style_header(ws2, 5)
    ws2.append(["Creatinine", "Enzymatic (IDMS-traceable)", "mg/dL", "CLIA", "15.0% or +/- 0.3 mg/dL"])
    ws2.append(["BUN", "Urease / glutamate dehydrogenase", "mg/dL", "CLIA", "9.0% or +/- 2 mg/dL"])
    auto_width(ws2)

    auto_width(ws)
    path = CHEM_DIR / "renal_function_qc.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# HEMATOLOGY
# ===================================================================

# -------------------------------------------------------------------
# 5. CBC QC on Sysmex XN-1000
# -------------------------------------------------------------------
def generate_cbc_qc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = [
        "Run Date", "Sample ID", "Parameter", "Level",
        "Value", "Unit", "Mean", "SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    dates = make_date_series("2025-08-01", 25, gap_days=1)

    parameters = [
        ("WBC", "x10^3/uL", [("L1 (Normal)", 7.5, 0.3), ("L2 (Abnormal)", 15.0, 0.5)]),
        ("RBC", "x10^6/uL", [("L1 (Normal)", 4.5, 0.08), ("L2 (Abnormal)", 3.0, 0.06)]),
        ("Hemoglobin", "g/dL", [("L1 (Normal)", 14.0, 0.2), ("L2 (Abnormal)", 10.0, 0.15)]),
        ("Hematocrit", "%", [("L1 (Normal)", 42.0, 0.6), ("L2 (Abnormal)", 30.0, 0.5)]),
        ("Platelets", "x10^3/uL", [("L1 (Normal)", 250.0, 8.0), ("L2 (Abnormal)", 100.0, 4.0)]),
    ]

    sample_counter = 0
    for run_i in range(25):
        for param_name, unit, levels in parameters:
            for level_name, mean, sd in levels:
                sample_counter += 1
                value = rand_val(mean, sd)
                ws.append([
                    dates[run_i],
                    f"CBC-QC-{sample_counter:04d}",
                    param_name,
                    level_name,
                    value,
                    unit,
                    mean,
                    sd,
                ])

    # Metadata sheet
    ws2 = wb.create_sheet("Instrument Info")
    ws2.append(["Parameter", "Value"])
    style_header(ws2, 2)
    ws2.append(["Instrument", "Sysmex XN-1000"])
    ws2.append(["QC Material", "Sysmex XN-CHECK L1 / L2"])
    ws2.append(["Runs", 25])
    ws2.append(["Levels", "L1 (Normal), L2 (Abnormal)"])
    ws2.append(["Parameters", "WBC, RBC, Hemoglobin, Hematocrit, Platelets"])
    auto_width(ws2)

    auto_width(ws)
    path = HEMA_DIR / "cbc_qc.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# -------------------------------------------------------------------
# 6. Coagulation QC on Sysmex CS-5100
# -------------------------------------------------------------------
def generate_coagulation_qc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = [
        "Run Date", "Sample ID", "Parameter", "Level",
        "Value", "Unit", "Mean", "SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    dates = make_date_series("2025-08-01", 20, gap_days=1)

    parameters = [
        ("PT", "sec", [("L1 (Normal)", 12.5, 0.3), ("L2 (Abnormal)", 25.0, 0.8)]),
        ("APTT", "sec", [("L1 (Normal)", 30.0, 1.0), ("L2 (Abnormal)", 55.0, 2.0)]),
        ("INR", "", [("L1 (Normal)", 1.0, 0.03), ("L2 (Abnormal)", 2.5, 0.08)]),
    ]

    sample_counter = 0
    for run_i in range(20):
        for param_name, unit, levels in parameters:
            for level_name, mean, sd in levels:
                sample_counter += 1
                value = rand_val(mean, sd)
                ws.append([
                    dates[run_i],
                    f"COAG-QC-{sample_counter:04d}",
                    param_name,
                    level_name,
                    value,
                    unit,
                    mean,
                    sd,
                ])

    # Metadata sheet
    ws2 = wb.create_sheet("Instrument Info")
    ws2.append(["Parameter", "Value"])
    style_header(ws2, 2)
    ws2.append(["Instrument", "Sysmex CS-5100"])
    ws2.append(["QC Material", "Dade Ci-Trol Coagulation Control Level 1 & 2"])
    ws2.append(["Runs", 20])
    ws2.append(["Levels", "L1 (Normal), L2 (Abnormal)"])
    ws2.append(["Parameters", "PT, APTT, INR"])
    ws2.append(["PT Reagent", "Dade Innovin (recombinant thromboplastin)"])
    ws2.append(["APTT Reagent", "Dade Actin FSL"])
    auto_width(ws2)

    auto_width(ws)
    path = HEMA_DIR / "coagulation_qc.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# IMMUNOASSAY
# ===================================================================

# -------------------------------------------------------------------
# 7. Thyroid QC on Cobas e801
# -------------------------------------------------------------------
def generate_thyroid_qc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = [
        "Run Date", "Sample ID", "Analyte", "Level",
        "Value", "Unit", "Mean", "SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    dates = make_date_series("2025-07-01", 20, gap_days=1)

    analytes = [
        ("TSH", "mIU/L", [("L1 (Normal)", 2.5, 0.15), ("L2 (Elevated)", 15.0, 0.8)]),
        ("FT4", "ng/dL", [("L1 (Normal)", 1.2, 0.06), ("L2 (Elevated)", 3.5, 0.15)]),
        ("FT3", "pg/mL", [("L1 (Normal)", 3.0, 0.15), ("L2 (Elevated)", 8.0, 0.35)]),
    ]

    sample_counter = 0
    for run_i in range(20):
        for analyte_name, unit, levels in analytes:
            for level_name, mean, sd in levels:
                sample_counter += 1
                value = rand_val(mean, sd)
                ws.append([
                    dates[run_i],
                    f"THY-QC-{sample_counter:04d}",
                    analyte_name,
                    level_name,
                    value,
                    unit,
                    mean,
                    sd,
                ])

    # Metadata sheet
    ws2 = wb.create_sheet("Assay Info")
    ws2.append(["Analyte", "Method", "Units", "TEa Source", "TEa%"])
    style_header(ws2, 5)
    ws2.append(["TSH", "Electrochemiluminescence (ECLIA)", "mIU/L", "RiliBaK", "23.7%"])
    ws2.append(["FT4", "Electrochemiluminescence (ECLIA)", "ng/dL", "RiliBaK", "18.3%"])
    ws2.append(["FT3", "Electrochemiluminescence (ECLIA)", "pg/mL", "Ricos BV", "12.2%"])
    auto_width(ws2)

    auto_width(ws)
    path = IMMU_DIR / "thyroid_qc.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# -------------------------------------------------------------------
# 8. Cardiac Markers QC
# -------------------------------------------------------------------
def generate_cardiac_markers_qc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = [
        "Run Date", "Sample ID", "Analyte", "Level",
        "Value", "Unit", "Mean", "SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    dates = make_date_series("2025-07-01", 15, gap_days=1)

    analytes = [
        ("Troponin I", "ng/mL", [("L1 (Normal)", 0.04, 0.005), ("L2 (Elevated)", 5.0, 0.25)]),
        ("CK-MB", "ng/mL", [("L1 (Normal)", 3.0, 0.2), ("L2 (Elevated)", 25.0, 1.2)]),
        ("BNP", "pg/mL", [("L1 (Normal)", 100.0, 8.0), ("L2 (Elevated)", 800.0, 40.0)]),
    ]

    sample_counter = 0
    for run_i in range(15):
        for analyte_name, unit, levels in analytes:
            for level_name, mean, sd in levels:
                sample_counter += 1
                value = rand_val(mean, sd)
                ws.append([
                    dates[run_i],
                    f"CARD-QC-{sample_counter:04d}",
                    analyte_name,
                    level_name,
                    value,
                    unit,
                    mean,
                    sd,
                ])

    # Metadata sheet
    ws2 = wb.create_sheet("Assay Info")
    ws2.append(["Analyte", "Method", "Units", "TEa Source", "TEa%"])
    style_header(ws2, 5)
    ws2.append(["Troponin I", "Chemiluminescence (high-sensitivity)", "ng/mL", "AACC/Ricos", "14.2%"])
    ws2.append(["CK-MB", "Chemiluminescence", "ng/mL", "CLIA", "30.0%"])
    ws2.append(["BNP", "Chemiluminescence", "pg/mL", "Ricos BV", "28.0%"])
    auto_width(ws2)

    auto_width(ws)
    path = IMMU_DIR / "cardiac_markers_qc.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# -------------------------------------------------------------------
# 9. Tumor Markers QC
# -------------------------------------------------------------------
def generate_tumor_markers_qc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Data"

    headers = [
        "Run Date", "Sample ID", "Analyte", "Level",
        "Value", "Unit", "Mean", "SD",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    dates = make_date_series("2025-07-01", 20, gap_days=1)

    analytes = [
        ("PSA", "ng/mL", [("L1 (Normal)", 2.0, 0.12), ("L2 (Elevated)", 15.0, 0.6)]),
        ("CEA", "ng/mL", [("L1 (Normal)", 2.5, 0.15), ("L2 (Elevated)", 20.0, 0.8)]),
        ("CA-125", "U/mL", [("L1 (Normal)", 15.0, 1.5), ("L2 (Elevated)", 200.0, 10.0)]),
        ("AFP", "ng/mL", [("L1 (Normal)", 5.0, 0.3), ("L2 (Elevated)", 100.0, 4.0)]),
    ]

    sample_counter = 0
    for run_i in range(20):
        for analyte_name, unit, levels in analytes:
            for level_name, mean, sd in levels:
                sample_counter += 1
                value = rand_val(mean, sd)
                ws.append([
                    dates[run_i],
                    f"TM-QC-{sample_counter:04d}",
                    analyte_name,
                    level_name,
                    value,
                    unit,
                    mean,
                    sd,
                ])

    # Metadata sheet
    ws2 = wb.create_sheet("Assay Info")
    ws2.append(["Analyte", "Method", "Units", "TEa Source", "TEa%"])
    style_header(ws2, 5)
    ws2.append(["PSA", "Electrochemiluminescence (ECLIA)", "ng/mL", "BV (Ricos)", "16.0%"])
    ws2.append(["CEA", "Electrochemiluminescence (ECLIA)", "ng/mL", "BV (Ricos)", "28.4%"])
    ws2.append(["CA-125", "Electrochemiluminescence (ECLIA)", "U/mL", "BV (Ricos)", "30.0%"])
    ws2.append(["AFP", "Electrochemiluminescence (ECLIA)", "ng/mL", "BV (Ricos)", "29.0%"])
    auto_width(ws2)

    auto_width(ws)
    path = IMMU_DIR / "tumor_markers_qc.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# SIGMA DATA - All Disciplines
# ===================================================================

def generate_sigma_all_disciplines():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sigma Inputs"

    headers = [
        "Assay", "Discipline", "TEa%", "Bias%", "CV%",
        "Expected Sigma", "Sigma Rating", "TEa Source",
        "Bias Source", "CV Source",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    def sigma_rating(sigma_val):
        if sigma_val >= 6.0:
            return "World Class"
        elif sigma_val >= 5.0:
            return "Excellent"
        elif sigma_val >= 4.0:
            return "Good"
        elif sigma_val >= 3.0:
            return "Marginal"
        elif sigma_val >= 2.0:
            return "Poor"
        else:
            return "Unacceptable"

    assays = [
        # (Assay, Discipline, TEa%, Bias%, CV%, TEa Source, Bias Source, CV Source)
        ("Glucose", "Clinical Chemistry", 6.96, 1.4, 1.8, "CLIA", "EQA 2024", "IQC 2024-Q4"),
        ("Creatinine", "Clinical Chemistry", 15.0, 3.2, 2.8, "CLIA", "EQA 2024", "IQC 2024-Q4"),
        ("Total Cholesterol", "Clinical Chemistry", 10.0, 2.0, 1.5, "NCEP", "EQA 2024", "IQC 2024-Q4"),
        ("HDL Cholesterol", "Clinical Chemistry", 13.0, 3.0, 2.5, "NCEP", "EQA 2024", "IQC 2024-Q4"),
        ("HbA1c", "Clinical Chemistry", 6.0, 1.0, 1.5, "NGSP", "EQA 2024", "IQC 2024-Q4"),
        ("Hemoglobin", "Hematology", 7.0, 1.0, 1.2, "CAP", "EQA 2024", "IQC 2024-Q4"),
        ("WBC", "Hematology", 15.0, 3.0, 3.5, "CAP", "EQA 2024", "IQC 2024-Q4"),
        ("Platelet Count", "Hematology", 25.0, 5.0, 4.0, "CAP", "EQA 2024", "IQC 2024-Q4"),
        ("PT (INR)", "Hematology", 15.0, 2.0, 2.5, "CLIA", "EQA 2024", "IQC 2024-Q4"),
        ("TSH", "Immunoassay", 23.7, 4.5, 3.2, "RiliBaK", "EQA 2024", "IQC 2024-Q4"),
        ("Troponin I", "Immunoassay", 14.2, 5.1, 4.0, "AACC", "EQA 2024", "IQC 2024-Q4"),
        ("PSA", "Immunoassay", 16.0, 3.0, 3.5, "BV (Ricos)", "EQA 2024", "IQC 2024-Q4"),
        ("SARS-CoV-2 Ct", "Molecular", 10.0, 2.0, 1.5, "Internal", "Method comparison", "IQC 2024-Q4"),
        ("HIV-1 VL", "Molecular", 15.0, 3.0, 2.5, "Internal", "VQA proficiency", "IQC 2024-Q4"),
    ]

    for assay_name, discipline, tea, bias, cv, tea_src, bias_src, cv_src in assays:
        sigma = round((tea - abs(bias)) / cv, 2)
        rating = sigma_rating(sigma)
        ws.append([
            assay_name, discipline, tea, bias, cv,
            sigma, rating, tea_src, bias_src, cv_src,
        ])

    # Sigma Scale Reference sheet
    ws2 = wb.create_sheet("Sigma Scale Reference")
    ws2.append(["Sigma Value", "Rating", "Defect Rate", "QC Strategy"])
    style_header(ws2, 4)
    ws2.append(["< 2.0", "Unacceptable", "> 308,537 DPM", "Cannot rely on QC alone; redesign method"])
    ws2.append(["2.0 - 3.0", "Poor", "66,807 - 308,537 DPM", "Maximum QC: multi-rule, N=4+"])
    ws2.append(["3.0 - 4.0", "Marginal", "6,210 - 66,807 DPM", "1-3s/2-2s/R-4s, N=2-4"])
    ws2.append(["4.0 - 5.0", "Good", "233 - 6,210 DPM", "1-3s, N=2"])
    ws2.append(["5.0 - 6.0", "Excellent", "3.4 - 233 DPM", "1-3s, N=1-2"])
    ws2.append(["> 6.0", "World Class", "< 3.4 DPM", "Minimal QC needed"])
    auto_width(ws2)

    # TEa sources sheet
    ws3 = wb.create_sheet("TEa Sources")
    ws3.append(["Source", "Description", "URL / Reference"])
    style_header(ws3, 3)
    ws3.append(["CLIA", "US Clinical Laboratory Improvement Amendments proficiency testing criteria",
                "42 CFR 493.931 - 493.937"])
    ws3.append(["NCEP", "National Cholesterol Education Program analytical goals",
                "NCEP ATP III Guidelines"])
    ws3.append(["NGSP", "National Glycohemoglobin Standardization Program",
                "http://www.ngsp.org"])
    ws3.append(["CAP", "College of American Pathologists proficiency testing limits",
                "https://www.cap.org"])
    ws3.append(["RiliBaK", "German Richtlinie der Bundesarztekammer",
                "https://www.bundesaerztekammer.de"])
    ws3.append(["AACC", "American Association for Clinical Chemistry / EFLM BV database",
                "https://biologicalvariation.eu"])
    ws3.append(["BV (Ricos)", "Biological variation desirable specifications (Ricos / EFLM)",
                "https://biologicalvariation.eu"])
    ws3.append(["Internal", "Laboratory-defined TEa based on clinical decision points",
                "N/A"])
    auto_width(ws3)

    auto_width(ws)
    path = SIGMA_DIR / "sigma_inputs_all_disciplines.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# VALIDATION DATA - Clinical Chemistry
# ===================================================================

# -------------------------------------------------------------------
# 10. Precision Study - Glucose (Intra/Inter-run)
# -------------------------------------------------------------------
def generate_precision_glucose():
    wb = openpyxl.Workbook()

    # Sheet 1: Intra-run precision (20 replicates in single run)
    ws1 = wb.active
    ws1.title = "Intra-Run Precision"

    headers = [
        "Sample ID", "Analyte", "Level", "Replicate",
        "Value (mg/dL)", "Run Date", "Instrument", "Operator",
    ]
    ws1.append(headers)
    style_header(ws1, len(headers))

    analyte = "Glucose"
    mean_val = 95.0
    sd_val = 2.5

    for rep in range(1, 21):
        value = rand_val(mean_val, sd_val)
        ws1.append([
            f"GLU-PREC-IR-{rep:03d}",
            analyte,
            "L1 (Normal)",
            rep,
            value,
            "2025-11-01",
            "Roche Cobas c501",
            "Tech-A",
        ])

    # Summary
    ws1.append([])
    ws1.append(["Summary Statistics"])
    ws1.append(["N Replicates", 20])
    ws1.append(["Target Mean", f"{mean_val} mg/dL"])
    ws1.append(["Target SD", f"{sd_val} mg/dL"])
    ws1.append(["Acceptance: CV%", "<3%"])
    auto_width(ws1)

    # Sheet 2: Inter-run precision (5 days, 4 replicates each)
    ws2 = wb.create_sheet("Inter-Run Precision")
    headers2 = [
        "Sample ID", "Analyte", "Level", "Run Day",
        "Run Date", "Replicate", "Value (mg/dL)", "Instrument", "Operator",
    ]
    ws2.append(headers2)
    style_header(ws2, len(headers2))

    dates = make_date_series("2025-11-01", 5, gap_days=1)
    operators = ["Tech-A", "Tech-B", "Tech-A", "Tech-C", "Tech-B"]

    sample_counter = 0
    for day_i in range(5):
        # Add slight day-to-day variation
        day_shift = random.gauss(0, 1.0)
        for rep in range(1, 5):
            sample_counter += 1
            value = rand_val(mean_val + day_shift, sd_val)
            ws2.append([
                f"GLU-PREC-ER-{sample_counter:03d}",
                analyte,
                "L1 (Normal)",
                day_i + 1,
                dates[day_i],
                rep,
                value,
                "Roche Cobas c501",
                operators[day_i],
            ])

    auto_width(ws2)

    path = VAL_DIR / "precision_glucose.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# -------------------------------------------------------------------
# 11. Linearity Study - Creatinine
# -------------------------------------------------------------------
def generate_linearity_creatinine():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Linearity Study"

    headers = [
        "Sample ID", "Analyte", "Nominal Concentration (mg/dL)",
        "Replicate", "Measured Value (mg/dL)",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    analyte = "Creatinine"
    levels = [0.5, 1.0, 2.0, 5.0, 10.0, 15.0, 20.0]

    # For realistic linearity data, measured = nominal * (1 + small_bias) + noise
    bias_factor = 1.02  # 2% proportional bias
    noise_sd_fraction = 0.015  # 1.5% CV

    sample_counter = 0
    for conc in levels:
        for rep in range(1, 4):
            sample_counter += 1
            noise = random.gauss(0, conc * noise_sd_fraction)
            measured = round(conc * bias_factor + noise, 3)
            # Ensure non-negative
            if measured < 0:
                measured = round(abs(noise), 3)
            ws.append([
                f"CREA-LIN-{sample_counter:03d}",
                analyte,
                conc,
                rep,
                measured,
            ])

    # Summary sheet
    ws2 = wb.create_sheet("Study Summary")
    ws2.append(["Parameter", "Value"])
    style_header(ws2, 2)
    ws2.append(["Analyte", analyte])
    ws2.append(["Instrument", "Roche Cobas c501"])
    ws2.append(["Method", "Enzymatic (IDMS-traceable)"])
    ws2.append(["Units", "mg/dL"])
    ws2.append(["N Levels", len(levels)])
    ws2.append(["Levels (mg/dL)", ", ".join(str(l) for l in levels)])
    ws2.append(["Replicates per Level", 3])
    ws2.append(["Total Measurements", sample_counter])
    ws2.append(["Expected R-squared", ">= 0.995"])
    ws2.append(["Acceptance", "Slope 0.95-1.05, R-sq >= 0.995, recovery 90-110%"])
    ws2.append(["Reportable Range", f"{levels[0]} - {levels[-1]} mg/dL"])
    auto_width(ws2)

    auto_width(ws)
    path = VAL_DIR / "linearity_creatinine.xlsx"
    wb.save(path)
    print(f"  Created: {path}")


# ===================================================================
# Main - Generate all files
# ===================================================================
def main():
    print("=" * 60)
    print("LabQC Multi-Discipline Example Data Generator")
    print("=" * 60)
    print()
    print(f"Output directory: {BASE_DIR}")
    print(f"Random seed: {SEED}")
    print()

    print("[Clinical Chemistry]")
    generate_glucose_qc()
    generate_lipid_panel_qc()
    generate_liver_function_qc()
    generate_renal_function_qc()
    print()

    print("[Hematology]")
    generate_cbc_qc()
    generate_coagulation_qc()
    print()

    print("[Immunoassay]")
    generate_thyroid_qc()
    generate_cardiac_markers_qc()
    generate_tumor_markers_qc()
    print()

    print("[Sigma Data - All Disciplines]")
    generate_sigma_all_disciplines()
    print()

    print("[Validation Data - Clinical Chemistry]")
    generate_precision_glucose()
    generate_linearity_creatinine()
    print()

    print("=" * 60)
    print("All multi-discipline files generated successfully.")
    print("=" * 60)


if __name__ == "__main__":
    main()
