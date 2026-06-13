"""Generic fallback parser for .xlsx QC files with smart column auto-detection."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from backend.parsers.base_parser import BaseParser


# Auto-detection patterns (tried in order, case-insensitive).
# Each list is ordered by specificity so more-specific names match first.
_VALUE_PATTERNS = [
    # English
    "ct value", "cq", "ct", "value", "result", "measurement",
    # Chinese
    "ct值", "cq值", "ct", "cq", "值", "结果", "测定结果", "浓度",
    "检测值", "测量值",
]
_LEVEL_PATTERNS = [
    # English
    "control level", "level", "sample name", "sample id", "sample",
    "control", "content",
    # Chinese
    "质控水平", "水平", "质控品", "级别", "样品名称", "样本名称",
    "样本", "样品", "质控级别", "质控等级",
]
_MEAN_PATTERNS = [
    # English
    "ct mean", "assigned mean", "target mean", "mean",
    # Chinese
    "均值", "靶值", "靶均值", "标准均值", "平均", "平均值",
    "目标均值", "设定均值",
]
_SD_PATTERNS = [
    # English
    "ct sd", "assigned sd", "target sd", "std dev", "sd",
    # Chinese
    "标准差", "标准偏差", "sd", "sd值",
]
_TARGET_PATTERNS = [
    # English
    "target name", "target", "analyte", "parameter", "test",
    # Chinese
    "靶标", "分析物", "项目", "检测项目", "参数", "指标",
]
_WELL_PATTERNS = [
    # English
    "well position", "well",
    # Chinese
    "孔位", "孔号", "孔", "位置",
]


def _find_column(headers_lower: list[str], patterns: list[str]) -> int | None:
    """Find the first header matching any pattern.

    Tries exact match first, then startswith.
    """
    # First try exact match
    for pattern in patterns:
        if pattern in headers_lower:
            return headers_lower.index(pattern)
    # Then try startswith (e.g. "value (mg/dl)" starts with "value")
    for pattern in patterns:
        for i, h in enumerate(headers_lower):
            if h.startswith(pattern):
                return i
    return None


class GenericParser(BaseParser):
    """Fallback parser that auto-detects columns or accepts a mapping config."""

    def can_handle(self, file_metadata: dict) -> bool:
        """Always returns True -- this is the last-resort fallback parser."""
        return True

    def parse(
        self,
        file_bytes: bytes,
        mapping_config: dict | None = None,
    ) -> dict:
        """Parse xlsx bytes using auto-detected columns or a user-provided mapping.

        Parameters
        ----------
        mapping_config : dict | None
            Maps canonical field names to actual column headers in the file.
            Expected keys: ``"value"`` (or ``"ct_value"``), ``"level"`` (or
            ``"control_level"``).
            Optional keys: ``"target"``, ``"well"``, ``"mean"``, ``"sd"``.
            If ``None``, columns are auto-detected from common header patterns.

        Returns
        -------
        dict
            ``{"rows": [{"control_level": str, "ct_value": float,
                          "target": str, "well": str,
                          "mean": float|None, "sd": float|None}, ...]}``
        """
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return {"rows": []}

        rows_iter = ws.iter_rows()
        header_row = next(rows_iter)
        headers = [
            str(c.value).strip() if c.value is not None else ""
            for c in header_row
        ]
        headers_lower = [h.lower() for h in headers]

        # Build column index map (original-cased header -> index)
        col_map: dict[str, int] = {}
        for idx, h in enumerate(headers):
            col_map[h] = idx

        # Resolve column indices: mapping_config overrides auto-detection
        value_idx = None
        level_idx = None
        mean_idx = None
        sd_idx = None
        target_idx = None
        well_idx = None

        if mapping_config:
            # Normalize keys: accept both "value"/"ct_value" and "level"/"control_level"
            mc = {k.lower(): v for k, v in mapping_config.items()}

            def _resolve(key_variants: list[str]) -> int | None:
                for k in key_variants:
                    header_name = mc.get(k)
                    if header_name:
                        # Find by exact match (case-insensitive)
                        hn_lower = header_name.lower()
                        for i, hl in enumerate(headers_lower):
                            if hl == hn_lower:
                                return i
                        # Try startswith
                        for i, hl in enumerate(headers_lower):
                            if hl.startswith(hn_lower):
                                return i
                return None

            value_idx = _resolve(["value", "ct_value"])
            level_idx = _resolve(["level", "control_level"])
            mean_idx = _resolve(["mean"])
            sd_idx = _resolve(["sd"])
            target_idx = _resolve(["target"])
            well_idx = _resolve(["well"])

        # Auto-detect any columns not resolved by mapping_config
        if value_idx is None:
            value_idx = _find_column(headers_lower, _VALUE_PATTERNS)
        if level_idx is None:
            level_idx = _find_column(headers_lower, _LEVEL_PATTERNS)
        if mean_idx is None:
            mean_idx = _find_column(headers_lower, _MEAN_PATTERNS)
        if sd_idx is None:
            sd_idx = _find_column(headers_lower, _SD_PATTERNS)
        if target_idx is None:
            target_idx = _find_column(headers_lower, _TARGET_PATTERNS)
        if well_idx is None:
            well_idx = _find_column(headers_lower, _WELL_PATTERNS)

        # Must have at least the value column to parse anything useful
        if value_idx is None:
            wb.close()
            return {"rows": []}

        parsed_rows: list[dict] = []
        for row in rows_iter:
            values = [c.value for c in row]

            # Value (the measured QC value -- Ct, Cq, concentration, count, etc.)
            ct_raw = values[value_idx] if value_idx < len(values) else None
            if ct_raw is None or str(ct_raw).strip().lower() in (
                "", "undetermined", "n/a", "nan",
            ):
                continue
            try:
                ct_value = float(ct_raw)
            except (ValueError, TypeError):
                continue

            # Control level
            level_raw = (
                values[level_idx] if level_idx is not None and level_idx < len(values) else None
            )
            control_level = _derive_control_level(
                str(level_raw).strip() if level_raw else ""
            )

            # Target / analyte (optional)
            target = ""
            if target_idx is not None and target_idx < len(values):
                tv = values[target_idx]
                target = str(tv).strip() if tv is not None else ""

            # Well (optional)
            well = ""
            if well_idx is not None and well_idx < len(values):
                wv = values[well_idx]
                well = str(wv).strip() if wv is not None else ""

            # Mean (optional)
            mean_val = None
            if mean_idx is not None and mean_idx < len(values):
                mean_raw = values[mean_idx]
                if mean_raw is not None and str(mean_raw).strip().lower() not in (
                    "", "undetermined", "n/a",
                ):
                    try:
                        mean_val = float(mean_raw)
                    except (ValueError, TypeError):
                        mean_val = None

            # SD (optional)
            sd_val = None
            if sd_idx is not None and sd_idx < len(values):
                sd_raw = values[sd_idx]
                if sd_raw is not None and str(sd_raw).strip().lower() not in (
                    "", "undetermined", "n/a",
                ):
                    try:
                        sd_val = float(sd_raw)
                    except (ValueError, TypeError):
                        sd_val = None

            parsed_rows.append(
                {
                    "control_level": control_level,
                    "ct_value": ct_value,
                    "mean": mean_val,
                    "sd": sd_val,
                    "target": target,
                    "well": well,
                }
            )

        wb.close()
        return {"rows": parsed_rows}

    def normalize_instrument_name(self) -> str:
        return "Generic"


def _derive_control_level(raw: str) -> str:
    """Best-effort mapping of a sample/level name to L1/L2/L3.

    Handles patterns from clinical chemistry, hematology, immunoassay,
    and molecular QC files.
    """
    if not raw:
        return "Unknown"

    lower = raw.lower()

    # Explicit level tags (most specific first) — English
    for tag in ("l1", "level 1", "level1", "水平1", "水平一", "低值"):
        if tag in lower:
            return "L1"
    for tag in ("l2", "level 2", "level2", "水平2", "水平二", "中值", "正常"):
        if tag in lower:
            return "L2"
    for tag in ("l3", "level 3", "level3", "水平3", "水平三", "高值"):
        if tag in lower:
            return "L3"

    # Clinical-chemistry qualitative tags — English
    if "low" in lower and "normal" not in lower:
        return "L1"
    if "normal" in lower:
        return "L1"
    if "mid" in lower or "medium" in lower:
        return "L2"
    if "high" in lower or "abnormal" in lower or "pathological" in lower or "elevated" in lower:
        return "L2"

    # Clinical-chemistry qualitative tags — Chinese
    if "低值" in lower and "正常" not in lower:
        return "L1"
    if "正常" in lower:
        return "L1"
    if "中值" in lower or "中等" in lower:
        return "L2"
    if "高值" in lower or "异常" in lower or "病理" in lower:
        return "L2"

    # Bio-Rad CFX "Pos Ctrl" / "Neg Ctrl" patterns
    if "neg" in lower:
        return "NTC"
    if "pos" in lower:
        # Try to extract a level from the Sample column (e.g. "QC-L1 Low Positive")
        return raw  # Keep as-is so upstream L1/L2 detection can work

    return raw or "Unknown"
