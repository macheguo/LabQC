"""Generate sample Excel files for testing LabQC end-to-end."""
import random
from pathlib import Path
from openpyxl import Workbook

SAMPLES_DIR = Path(__file__).parent


def generate_qc_quantstudio():
    """Generate a QuantStudio-style QC export with realistic Ct values."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # QuantStudio header columns
    headers = ["Well", "Well Position", "Sample Name", "Target Name", "CT", "Ct Mean", "Ct SD"]
    ws.append(headers)

    # Generate 20 QC data points for a single assay
    # Simulate L1 (low positive) and L2 (high positive) controls
    random.seed(42)  # Reproducible
    mean_l1, sd_l1 = 28.0, 0.5
    mean_l2, sd_l2 = 22.0, 0.4

    well_num = 1
    for run in range(10):
        # L1 control
        ct = round(random.gauss(mean_l1, sd_l1), 2)
        ws.append([well_num, f"A{well_num}", f"QC-L1-Run{run+1}", "SARS-CoV-2", ct, mean_l1, sd_l1])
        well_num += 1

        # L2 control
        ct = round(random.gauss(mean_l2, sd_l2), 2)
        ws.append([well_num, f"B{well_num}", f"QC-L2-Run{run+1}", "SARS-CoV-2", ct, mean_l2, sd_l2])
        well_num += 1

    path = SAMPLES_DIR / "sample_qc_quantstudio.xlsx"
    wb.save(path)
    print(f"Created {path}")


def generate_qc_with_violations():
    """Generate QC data that will trigger Westgard violations."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["Well", "Well Position", "Sample Name", "Target Name", "CT", "Ct Mean", "Ct SD"]
    ws.append(headers)

    mean, sd_val = 25.0, 1.0
    well_num = 1

    # Normal points first
    for i in range(5):
        ct = round(mean + random.uniform(-0.5, 0.5), 2)
        ws.append([well_num, f"A{well_num}", f"QC-L1-Run{i+1}", "HIV-1", ct, mean, sd_val])
        well_num += 1

    # 1-3s violation: point beyond 3 SD
    ct = round(mean + 3.2 * sd_val, 2)
    ws.append([well_num, f"A{well_num}", "QC-L1-Run6", "HIV-1", ct, mean, sd_val])
    well_num += 1

    # More normal points
    for i in range(4):
        ct = round(mean + random.uniform(-0.3, 0.3), 2)
        ws.append([well_num, f"A{well_num}", f"QC-L1-Run{7+i}", "HIV-1", ct, mean, sd_val])
        well_num += 1

    path = SAMPLES_DIR / "sample_qc_violations.xlsx"
    wb.save(path)
    print(f"Created {path}")


def generate_validation_lod():
    """Generate LOD validation dataset."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LOD Data"

    headers = ["Run", "Replicate", "Concentration", "Measured Value"]
    ws.append(headers)

    random.seed(123)
    # Blank replicates
    for rep in range(1, 21):
        value = round(random.gauss(50, 5), 2)
        ws.append([1, rep, 0, value])

    path = SAMPLES_DIR / "sample_validation_lod.xlsx"
    wb.save(path)
    print(f"Created {path}")


def generate_validation_linearity():
    """Generate linearity validation dataset."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Linearity Data"

    headers = ["Level", "Expected", "Measured"]
    ws.append(headers)

    random.seed(456)
    # 5 concentration levels with slight scatter
    for level, expected in enumerate([10, 50, 100, 500, 1000], 1):
        measured = round(expected * 1.02 + random.gauss(0, expected * 0.03), 2)
        ws.append([level, expected, measured])

    path = SAMPLES_DIR / "sample_validation_linearity.xlsx"
    wb.save(path)
    print(f"Created {path}")


def generate_sigma_reference():
    """Generate a reference file with sample Sigma inputs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sigma Inputs"

    headers = ["Assay", "TEa %", "Bias %", "CV %"]
    ws.append(headers)

    data = [
        ["Glucose", 10.0, 1.5, 1.2],
        ["Creatinine", 15.0, 3.0, 2.5],
        ["HbA1c", 6.0, 1.0, 1.5],
        ["TSH", 25.0, 5.0, 3.0],
        ["Troponin I", 20.0, 8.0, 4.5],
    ]
    for row in data:
        ws.append(row)

    path = SAMPLES_DIR / "sample_sigma_inputs.xlsx"
    wb.save(path)
    print(f"Created {path}")


if __name__ == "__main__":
    SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
    generate_qc_quantstudio()
    generate_qc_with_violations()
    generate_validation_lod()
    generate_validation_linearity()
    generate_sigma_reference()
    print("All sample files generated.")
